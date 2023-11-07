import logging
import sys
import argparse
from openpyxl import Workbook
import clr

# Add the necessary references
clr.AddReference("System")
clr.AddReference(r"C:\Program Files\Audio Precision\APx500 8.0\API\AudioPrecision.API2.dll")
clr.AddReference(r"C:\Program Files\Audio Precision\APx500 8.0\API\AudioPrecision.API.dll")

from AudioPrecision.API import *

# Setup Logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Argument Parsing
parser = argparse.ArgumentParser(description='APx Automation Script')
parser.add_argument('-name', '--filename', type=str, help='Name of the Excel file', required=True)
args = parser.parse_args()

# Main Class (example name: APxAutomation)
class APxAutomation:
    def __init__(self):
        self.APx = APx500_Application()
        self.checkedData = []

    def GetCheckedData(self, sender=None, args=None):

        if not self.APx:
            logging.error("APx is None. Launch AP Software first.")
            return

        checked_signal_paths = []
        try:
            for sp_idx, sp in enumerate(self.APx.Sequence):
                signal_path = ISignalPath(sp)
                #logging.info(f"Processing Signal Path: {signal_path.Name}")
                if not signal_path.Checked:
                    continue
                current_sp = {"name": signal_path.Name, "measurements": []}
                #logging.info(f"Checked Signal Path: {signal_path.Name}")
                sp_failed, sp_error = False, False  # added sp_error to track signal paths with errors

                for m_idx, m in enumerate(signal_path):
                    measurement = ISequenceMeasurement(m)
                    if not measurement.Checked:
                        continue
                    current_measurement = {"name": measurement.Name, "results": []}
                    #logging.info(f"\tChecked Measurement: {measurement.Name}")

                    for result_idx, result in enumerate(measurement.SequenceResults):
                        sequence_result = ISequenceResult(result)
                        failed = self.is_result_failed(signal_path.Name, measurement.Name, sequence_result.Name)
                        error = self.result_error(signal_path.Name, measurement.Name, sequence_result.Name)

                        status = "ERROR" if error else ("FAIL" if failed else "PASS")
                        logging.info(f"{signal_path.Name} | {measurement.Name} | {sequence_result.Name} -> {status}")

                        current_result = {
                            'name': sequence_result.Name,
                            'result_object': sequence_result,
                            'data': self.process_measurement_data(sequence_result),
                            'passed': not (failed or error)
                        }
                        current_measurement["results"].append(current_result)

                    current_sp["measurements"].append(current_measurement)

                checked_signal_paths.append(current_sp)
                #logging.info(f"Added a new signal path to checked_signal_paths. Total items now: {len(checked_signal_paths)}")

        except Exception as e:
            logging.error(f"An unexpected error occurred: {e}\nCheck the log file for more details.")

        return checked_signal_paths
    
    def process_measurement_data(self, sequence_result):
        data = {}
        
        # Check if the result object has the 'PassedLimitChecks' attribute
        if hasattr(sequence_result, 'PassedLimitChecks'):
            passed_limit_checks = sequence_result.PassedLimitChecks
            logging.info(f"\t\tPassedLimitChecks for {sequence_result.Name}: {passed_limit_checks}")
        else:
            passed_limit_checks = False
            #logging.warning(f"\t\t{sequence_result.Name} does not have PassedLimitChecks attribute. Setting to False.")

        # Check and Get XYValues
        for vertical_axis in [VerticalAxis.Left, VerticalAxis.Right]:
            channel_count = sequence_result.GetXYChannelCount(vertical_axis)
            if channel_count > 0:
                xValues, yValues = [], []
                for ch in range(channel_count):
                    graphPoints = sequence_result.GetXYValues(ch, vertical_axis, SourceDataType.Measured, 1)
                    if graphPoints:
                        xVals, yVals = [point.X for point in graphPoints], [point.Y for point in graphPoints]
                        xValues.append(xVals)
                        yValues.append(yVals)
                data['xValues'] = xValues
                data['yValues'] = yValues
        
        # Check for meter values
        if sequence_result.HasMeterValues:
            meterValues = sequence_result.GetMeterValues()
            data['meterValues'] = meterValues
            #logging.info(f"\t\tFound Meter Values for Result: {sequence_result.Name}")
            # Log each meter value directly.
            #for idx, value in enumerate(meterValues):
            #    logging.info(f"\t\t\tMeter Value {idx}: {value}")

        # Check and Get RawTextResults
        if sequence_result.HasRawTextResults:
            rawTextResults = sequence_result.GetRawTextResults()
            data['rawTextResults'] = rawTextResults
            logging.info(f"\t\tFound Raw Text Results for Result: {sequence_result.Name}")

        # Check and Get XValues and YValues
        for ch in range(sequence_result.GetXYChannelCount(VerticalAxis.Left)):
            for vertical_axis in [VerticalAxis.Left, VerticalAxis.Right]:
                for data_type in [SourceDataType.Measured]:  # Only considering SourceDataType.Measured as per the older logic
                    if hasattr(sequence_result, 'HasXValues') and sequence_result.HasXValues(ch, vertical_axis, data_type):
                        xValues = sequence_result.GetXValues(ch, vertical_axis, data_type, 1)
                        data[f'xValues_{vertical_axis}_{data_type}'] = xValues
                    if hasattr(sequence_result, 'HasYValues') and sequence_result.HasYValues(ch, vertical_axis, data_type):
                        yValues = sequence_result.GetYValues(ch, vertical_axis, data_type, 1)
                        data[f'yValues_{vertical_axis}_{data_type}'] = yValues

        return data

    def ExportCheckedDataToExcel(self, checkedData, filename):

        self.export_to_excel(checkedData, filename)

    @staticmethod
    def sanitize_sheet_name(name):
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '-')
        return name[:31]

    @staticmethod
    def abbreviate_name(name, max_length=10):
        if len(name) <= max_length:
            return name
        words = name.split(' ')
        if len(words) == 1:
            return name[:max_length]
        abbreviation = ''.join([word[0] for word in words])
        if len(abbreviation) <= max_length:
            return abbreviation
        return abbreviation[:max_length]
    
    def unique_sheet_name(self, workbook, base_name):
        count = 2
        sheet_name = base_name
        while sheet_name in workbook.sheetnames:
            sheet_name = f"{base_name} ({count})"
            count += 1
        return sheet_name
    
    def is_result_failed(self, signal_path_name, measurement_name, result_name):
        try:
            for idx in range(self.APx.Sequence.FailedResults.Count):
                failed_result = self.APx.Sequence.FailedResults[idx]
                if (failed_result.Name == result_name and
                    failed_result.MeasurementName == measurement_name and
                    failed_result.SignalPathName == signal_path_name):
                    return not failed_result.PassedResult
        except Exception as e:
            logging.error(f"Error retrieving failed status for result {result_name}: {str(e)}")
        return False
    
    def result_error(self, signal_path_name, measurement_name, result_name):
        try:
            for idx in range(self.APx.Sequence.Results.Count):
                result = self.APx.Sequence.Results[idx]
                if (result.Name == result_name and
                    result.MeasurementName == measurement_name and
                    result.SignalPathName == signal_path_name):
                    return result.HasErrorMessage
        except Exception as e:
            logging.error(f"Error retrieving error status for result {result_name}: {str(e)}")
        return False

    def export_to_excel(self, checked_signal_paths, unit_descriptor=None, args=None):
        wb = Workbook()
        ws = wb.active
        has_created_sheet = False

        try:
            # unit_no = self.unitInput.Text.strip()
            if unit_descriptor:  # Incorporate the descriptor into the filename
                file_name = f"{unit_descriptor}.xlsx" if unit_descriptor else "exported_data.xlsx"

            # Define cleaned_descriptor to remove the suffixes
            cleaned_descriptor = unit_descriptor.replace("_PASS", "").replace("_FAIL", "").replace("_ERRORS", "")

            for sp in checked_signal_paths:
                for measurement in sp["measurements"]:
                    for result in measurement["results"]:

                        xy_values_found = False  # Add a flag to track if xy values are found

                        # Handle xy values
                        if 'xValues' in result['data'] and 'yValues' in result['data']:
                            xValues = result['data']['xValues']
                            yValues = result['data']['yValues']
                            
                            if not xValues or not yValues:  # Check if either xValues or yValues is empty
                                logging.warning(f"xValues or yValues is empty for {result['name']}.")
                                continue
                            
                            if isinstance(xValues[0], (list, tuple)):
                                xValues = xValues[0]
                                
                            if not isinstance(yValues[0], (list, tuple)):
                                yValues = [yValues]
                            
                            sheet_title = f"{self.abbreviate_name(sp['name'])}_{self.abbreviate_name(measurement['name'])}_{self.abbreviate_name(result['name'])}"
                            sheet_name = self.unique_sheet_name(wb, sheet_title)
                            ws = wb.create_sheet(title=sheet_name)
                            ws.append([f'Signal Path: {sp["name"]}'])
                            ws.append([f'Measurement: {measurement["name"]}'])
                            ws.append([f'Result: {result["name"]}'])
                            # Modified header to include Serial Number for Y Values
                            cleaned_descriptor = unit_descriptor.replace("_PASS", "").replace("_FAIL", "").replace("_ERRORS", "")
                            ws.append(["X Values"] + [f"{cleaned_descriptor} Ch{idx+1}" for idx in range(len(result['data']['yValues']))])
                            if "ch2" in result['name'].lower():
                                continue  # skip the ch2 result to avoid duplicating data
                        
                            # Writing to Excel
                            for row_idx, xValue in enumerate(xValues):
                                yRow = []
                                for ch_idx, yValueSet in enumerate(result['data']['yValues']):
                                    if not isinstance(yValueSet, (list, tuple)):
                                        logging.error(f"Invalid yValue data structure at yValues channel index {ch_idx}. Expected list or tuple, but got {type(yValueSet)}.")
                                        yValueSet = [yValueSet]  # Convert single float to a list for consistency.
                                    
                                    try:
                                        yRow.append(yValueSet[row_idx])
                                    except IndexError:
                                        logging.warning(f"Index {row_idx} out of bounds for yValues set at channel index {ch_idx}.")
                                        yRow.append(None)
                                row = [xValue] + yRow
                                ws.append(row)
                                has_created_sheet = True 
                                xy_values_found = True  # Set the flag here!


                        if 'meterValues' in result['data']:
                            # Create a new sheet for Meter Values.
                            meter_sheet_title = f"{self.abbreviate_name(sp['name'])}_{self.abbreviate_name(measurement['name'])}_{self.abbreviate_name(result['name'])}"
                            meter_sheet_name = self.unique_sheet_name(wb, meter_sheet_title)
                            meter_ws = wb.create_sheet(title=meter_sheet_name)
                            
                            meter_ws.append([f'Signal Path: {sp["name"]}'])
                            meter_ws.append([f'Measurement: {measurement["name"]}'])
                            meter_ws.append([f'Result: {result["name"]}'])
                            
                            # Add the headers for 'Channels' and the serial number
                            meter_ws.append(['Channels', cleaned_descriptor])
                            
                            meterValues = result['data']['meterValues']
                            
                            # Adjusted to add the channel name to the start of each row
                            for idx, val in enumerate(meterValues):
                                meter_ws.append([f"Ch{idx+1}", val])  # The channel name "Chx" and its corresponding value on the same row
                                                        
                            has_created_sheet = True  # Set has_created_sheet to True when creating a sheet for meterValues
                                        
                        if 'rawTextResults' in result['data']:
                            ws.append(["Raw Text Results"])
                            rawTextResults = result['data']['rawTextResults']
                            ws.append(rawTextResults)
                            has_created_sheet = True

                        for vertical_axis in [VerticalAxis.Left, VerticalAxis.Right]:
                            for data_type_str in ["Measured", "Fitted", "Residual"]:
                                if hasattr(SourceDataType, data_type_str):
                                    data_type = getattr(SourceDataType, data_type_str)
                                x_key = f'xValues_{vertical_axis}_{data_type}'
                                y_key = f'yValues_{vertical_axis}_{data_type}'
                                if x_key in result['data'] and y_key in result['data']:
                                    ws.append([f"X Values ({vertical_axis}, {data_type})"] + [f"Y Values CH{idx+1} ({vertical_axis}, {data_type})" for idx in range(len(result['data'][y_key]))])
                                    for row_idx, xValue in enumerate(result['data'][x_key]):
                                        row = [xValue] + [result['data'][y_key][ch_idx][row_idx] for ch_idx in range(len(result['data'][y_key]))]
                                        ws.append(row)
                                    xy_values_found = True  # Update the flag if xy values are found

                        #if not xy_values_found:  # This remains outside the nested loop
                        #    logging.warning(f"{result['name']} does not have xy values.")

            if has_created_sheet:
                # Remove the default "Sheet" if it exists in the workbook.
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
                wb.save(file_name)
                logging.info(f"Exported data to {file_name}.")
            else:
                logging.warning("No data to export.")
        except Exception as e:
            logging.exception("An error occurred during Excel export:")
            logging.error(f"An unexpected error occurred: {e}\nCheck the log file for more details.")


# Main execution
if __name__ == "__main__":
    try:
        apx_auto = APxAutomation()
        checked_data = apx_auto.GetCheckedData()
        if checked_data:
            apx_auto.ExportCheckedDataToExcel(checked_data, args.filename)
        else:
            logging.warning("No checked data available for export.")
    except Exception as e:
        logging.error(f"An error occurred in the APx Automation script: {e}")
        sys.exit(1)
