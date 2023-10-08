"""
Author: Joshua Levy
Date: September 29, 2023
Description: A sample script to interface with the Audio Precision (APx) API.
This script provides a GUI with a button to fetch the results data from the APx software for all checked signal paths, 
measurements, and results.
"""

import clr
from openpyxl import Workbook, load_workbook
import logging
import os
import copy
import tkinter as tk
from tkinter import filedialog
import datetime


# Add the necessary references
clr.AddReference("System")  # Add reference to the System assembly
clr.AddReference("System.Drawing")
clr.AddReference("System.Windows.Forms")
clr.AddReference(r"C:\Program Files\Audio Precision\APx500 8.0\API\AudioPrecision.API2.dll")
clr.AddReference(r"C:\Program Files\Audio Precision\APx500 8.0\API\AudioPrecision.API.dll")

from System import Array
from System.Drawing import Point, Size, Color, Font, FontStyle
from System.Windows.Forms import Application, Form, TextBox, Label, Button, ListBox, SelectionMode
from System.Windows.Forms import View, ListView, HorizontalAlignment, ListViewItem, GroupBox
from System.Windows.Forms import ScrollBars, CheckBox, OpenFileDialog, DialogResult
from AudioPrecision.API import *

# Configure logging
logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s %(levelname)-8s %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S',
                    filename='apx_script.log',
                    filemode='w')

console = logging.StreamHandler()
console.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(name)-12s: %(levelname)-8s %(message)s')
console.setFormatter(formatter)
logging.getLogger('').addHandler(console)


class APxContainer(Form):

    PASS_COLOR = Color.Green
    FAIL_COLOR = Color.Red
    ERROR_COLOR = Color.Orange


    def __init__(self):
        Form.__init__(self)
        self.Text = 'APx Data Extractor v0.10'
        self.Height = 650  # Adjusted the height to accommodate the fourth ListBox below the initial three
        self.Width = 1000

        # Unit Input TextBox and Label
        self.unitLabel = Label()
        self.unitLabel.Text = "Serial Number"
        self.unitLabel.Location = Point(20, 20)
        self.unitLabel.Size = Size(80, 20)
        self.unitLabel.Font = Font(self.unitLabel.Font, FontStyle.Bold)
        self.Controls.Add(self.unitLabel)

        self.unitInput = TextBox()
        self.unitInput.Location = Point(100, 20)
        self.unitInput.Size = Size(200, 20)
        self.Controls.Add(self.unitInput)


        # Buttons
        self.bRunSequence = Button()
        self.bRunSequence.Text = "Run Sequence"
        self.bRunSequence.Location = Point(40, 30)
        self.bRunSequence.Size = Size(200, 30)
        self.bRunSequence.Click += self.APxRunSequence
        self.Controls.Add(self.bRunSequence)

        self.bRetrieveCheckedData = Button()
        self.bRetrieveCheckedData.Text = "Retrieve Checked Data"
        self.bRetrieveCheckedData.Location = Point(40, 70)
        self.bRetrieveCheckedData.Size = Size(200, 30)
        self.bRetrieveCheckedData.Click += self.GetCheckedData
        self.Controls.Add(self.bRetrieveCheckedData)

        self.bExportToExcel = Button()
        self.bExportToExcel.Text = "Export All Checked Results"
        self.bExportToExcel.Location = Point(40, 110)
        self.bExportToExcel.Size = Size(200, 30)
        self.bExportToExcel.Click += self.ExportCheckedDataToExcel
        self.Controls.Add(self.bExportToExcel)

        # "Select File" Button
        self.bSelectFile = Button()
        self.bSelectFile.Text = "Select File"
        self.bSelectFile.Location = Point(140, 150)  # Adjust the location accordingly
        self.bSelectFile.Size = Size(100, 30)
        self.bSelectFile.Click += lambda sender, args: self.select_file(sender, args, "All")
        self.bSelectFile.Enabled = False  # Initially greyed-out
        self.Controls.Add(self.bSelectFile)

        # "Append to" Checkbox
        self.appendCheckbox = CheckBox()
        self.appendCheckbox.Text = "Append to"
        self.appendCheckbox.Location = Point(50, 150)  # You may adjust the location based on other components
        self.appendCheckbox.CheckedChanged += self.toggleSelectFileButton
        self.Controls.Add(self.appendCheckbox)
        
        # ListViews
        self.checkedSignalPathsList = ListView()
        self.checkedSignalPathsList.View = View.Details  # Set to Details
        self.checkedSignalPathsList.Location = Point(320, 20)
        self.checkedSignalPathsList.Size = Size(200, 110)
        self.checkedSignalPathsList.SelectedIndexChanged += self.updateCheckedMeasurementsList
        self.checkedSignalPathsList.Columns.Add('Signal Path', -2, HorizontalAlignment.Left)  # Add appropriate column
        self.Controls.Add(self.checkedSignalPathsList)

        self.checkedMeasurementsList = ListView()
        self.checkedMeasurementsList.View = View.Details
        self.checkedMeasurementsList.Location = Point(540, 20)
        self.checkedMeasurementsList.Size = Size(200, 110)
        self.checkedMeasurementsList.SelectedIndexChanged += self.updateCheckedResultsList
        self.checkedMeasurementsList.Columns.Add('Measurement', -2, HorizontalAlignment.Left)
        self.Controls.Add(self.checkedMeasurementsList)

        self.checkedResultsList = ListView()
        self.checkedResultsList.View = View.Details
        self.checkedResultsList.Location = Point(760, 20)
        self.checkedResultsList.Size = Size(200, 110)
        self.checkedResultsList.Columns.Add('Result', -2, HorizontalAlignment.Left)
        self.Controls.Add(self.checkedResultsList)

        # ListBox
        self.selectedResultsList = ListBox()
        self.selectedResultsList.Location = Point(320, 150)
        self.selectedResultsList.Size = Size(640, 140)
        self.Controls.Add(self.selectedResultsList)

        # More buttons

        self.bPinSelectedSignalPath = Button()
        self.bPinSelectedSignalPath.Text = "Pin Selected Signal Path"
        self.bPinSelectedSignalPath.Location = Point(40, 190)
        self.bPinSelectedSignalPath.Size = Size(200, 30)
        self.bPinSelectedSignalPath.Click += self.PinSelectedSignalPath
        self.Controls.Add(self.bPinSelectedSignalPath)

        self.bPinSelectedMeasurement = Button()
        self.bPinSelectedMeasurement.Text = "Pin Selected Measurement"
        self.bPinSelectedMeasurement.Location = Point(40, 230)  # Adjust the location as necessary
        self.bPinSelectedMeasurement.Size = Size(200, 30)
        self.bPinSelectedMeasurement.Click += self.PinSelectedMeasurement
        self.Controls.Add(self.bPinSelectedMeasurement)

        self.bAddSelectedResult = Button()
        self.bAddSelectedResult.Text = "Pin Selected Result"
        self.bAddSelectedResult.Location = Point(40, 270)
        self.bAddSelectedResult.Size = Size(200, 30)
        self.bAddSelectedResult.Click += self.AddSelectedResult
        self.Controls.Add(self.bAddSelectedResult)

        self.bRerunPinnedTests = Button()
        self.bRerunPinnedTests.Text = "Rerun Pinned Tests"
        self.bRerunPinnedTests.Location = Point(40, 310)
        self.bRerunPinnedTests.Size = Size(200, 30)
        self.bRerunPinnedTests.Click += self.RerunPinnedTests
        self.Controls.Add(self.bRerunPinnedTests)

        self.bExportSelectedResults = Button()
        self.bExportSelectedResults.Text = "Export Selected Results"
        self.bExportSelectedResults.Location = Point(40, 350)
        self.bExportSelectedResults.Size = Size(200, 30)
        self.bExportSelectedResults.Click += self.ExportSelectedResults
        self.Controls.Add(self.bExportSelectedResults)

        self.bClearAll = Button()
        self.bClearAll.Text = "Clear All"
        self.bClearAll.Location = Point(40, 390)  # Adjust the location as needed
        self.bClearAll.Size = Size(200, 30)
        self.bClearAll.Click += self.ClearAll
        self.Controls.Add(self.bClearAll)

        self.controlButtonFrame = GroupBox()
        self.controlButtonFrame.Text = "Control Buttons"
        self.controlButtonFrame.Location = Point(20, 60)  # Adjust as necessary
        self.controlButtonFrame.Size = Size(280, 440)  # Adjust the size to encompass all buttons

        self.Controls.Add(self.controlButtonFrame)
        # Now, add all the buttons to the frame:
        self.controlButtonFrame.Controls.Add(self.bRunSequence)
        self.controlButtonFrame.Controls.Add(self.bRetrieveCheckedData)
        self.controlButtonFrame.Controls.Add(self.bExportToExcel)
        self.controlButtonFrame.Controls.Add(self.bExportToExcel)
        self.controlButtonFrame.Controls.Add(self.bSelectFile)
        self.controlButtonFrame.Controls.Add(self.appendCheckbox)
        self.controlButtonFrame.Controls.Add(self.bAddSelectedResult)
        self.controlButtonFrame.Controls.Add(self.bPinSelectedSignalPath)
        self.controlButtonFrame.Controls.Add(self.bPinSelectedMeasurement)
        self.controlButtonFrame.Controls.Add(self.bRerunPinnedTests)
        self.controlButtonFrame.Controls.Add(self.bExportSelectedResults)
        self.controlButtonFrame.Controls.Add(self.bClearAll)

        # Creating GroupBox for Test Result Statistics
        self.statisticsGroupBox = GroupBox()
        self.statisticsGroupBox.Text = "Test Result Statistics"
        self.statisticsGroupBox.Location = Point(310, 300)  # Adjust as needed
        self.statisticsGroupBox.Size = Size(680, 300)  # Adjust as needed

        self.total_tests_cumulative = 0
        self.total_passed_cumulative = 0
        self.total_failed_cumulative = 0
        self.total_errors_cumulative = 0

        # Labels for Heading
        headings = ["Passing", "Fail", "Error"]
        for i, text in enumerate(headings):
            label = Label()
            label.Text = text
            label.Location = Point(10 + 220 * i, 20)
            label.Size = Size(200, 20)
            self.statisticsGroupBox.Controls.Add(label)

        # Number Of Items Labels
        self.totalItemsLabel = Label()
        self.totalItemsLabel.Text = "Total # of Tests: N/A"
        self.totalItemsLabel.Location = Point(10, 50)
        self.totalItemsLabel.Size = Size(200, 20)
        self.statisticsGroupBox.Controls.Add(self.totalItemsLabel)

        # For displaying Number of Passed, Failed and Errored items
        for i in range(3):
            label = Label()
            label.Text = "N/A"
            label.Location = Point(10 + 220 * i, 50)
            label.Size = Size(200, 20)
            self.statisticsGroupBox.Controls.Add(label)
            if i == 0:
                self.passedNumberLabel = label
            elif i == 1:
                self.failedNumberLabel = label
            else:
                self.errorNumberLabel = label

        # Percentage Labels
        self.passRateLabel = Label()
        self.passRateLabel.Text = "Pass Rate: N/A"
        self.passRateLabel.Location = Point(10, 80)
        self.passRateLabel.Size = Size(200, 20)
        self.statisticsGroupBox.Controls.Add(self.passRateLabel)

        self.failRateLabel = Label()
        self.failRateLabel.Text = "Fail Rate: N/A"
        self.failRateLabel.Location = Point(230, 80)
        self.failRateLabel.Size = Size(200, 20)
        self.statisticsGroupBox.Controls.Add(self.failRateLabel)

        self.errorRateLabel = Label()
        self.errorRateLabel.Text = "Error Rate: N/A"
        self.errorRateLabel.Location = Point(450, 80)
        self.errorRateLabel.Size = Size(200, 20)
        self.statisticsGroupBox.Controls.Add(self.errorRateLabel)

        # List of Passing, Failing, and Errors
        self.passedListBox = TextBox()
        self.passedListBox.Multiline = True
        self.passedListBox.ScrollBars = ScrollBars.Vertical
        self.passedListBox.Text = "List of Passing: N/A"
        self.passedListBox.Location = Point(10, 110)
        self.passedListBox.Size = Size(200, 100)
        self.statisticsGroupBox.Controls.Add(self.passedListBox)

        self.failureListTextBox = TextBox()
        self.failureListTextBox.Multiline = True
        self.failureListTextBox.ScrollBars = ScrollBars.Vertical
        self.failureListTextBox.Text = "List of Failures: N/A"
        self.failureListTextBox.Location = Point(230, 110)
        self.failureListTextBox.Size = Size(200, 100)
        self.statisticsGroupBox.Controls.Add(self.failureListTextBox)

        self.errorListTextBox = TextBox()
        self.errorListTextBox.Multiline = True
        self.errorListTextBox.ScrollBars = ScrollBars.Vertical
        self.errorListTextBox.Text = "List of Errors: N/A"
        self.errorListTextBox.Location = Point(450, 110)
        self.errorListTextBox.Size = Size(200, 100)
        self.statisticsGroupBox.Controls.Add(self.errorListTextBox)

        # Add the GroupBox to the main form's controls
        self.Controls.Add(self.statisticsGroupBox)

        # Add export buttons below the list boxes
        self.selectedPassFilePath = None
        self.selectedFailFilePath = None
        self.selectedErrorFilePath = None

        self.exportPassButton = Button()
        self.exportPassButton.Text = "Export Passed"
        self.exportPassButton.Location = Point(10, 220)
        self.exportPassButton.Size = Size(90, 25)
        self.exportPassButton.Click += self.export_pass
        self.statisticsGroupBox.Controls.Add(self.exportPassButton)

        self.exportFailButton = Button()
        self.exportFailButton.Text = "Export Failures"
        self.exportFailButton.Location = Point(230, 220)
        self.exportFailButton.Size = Size(90, 25)
        self.exportFailButton.Click += self.export_fail
        self.statisticsGroupBox.Controls.Add(self.exportFailButton)

        self.exportErrorButton = Button()
        self.exportErrorButton.Text = "Export Errors"
        self.exportErrorButton.Location = Point(450, 220)
        self.exportErrorButton.Size = Size(90, 25)
        self.exportErrorButton.Click += self.export_error
        self.statisticsGroupBox.Controls.Add(self.exportErrorButton)

        # Add append to file checkbox for export passed
        self.appendPassCheckBox = CheckBox()
        self.appendPassCheckBox.Text = "Append to file"
        self.appendPassCheckBox.Location = Point(110, 220)
        self.appendPassCheckBox.Size = Size(100, 25)
        self.statisticsGroupBox.Controls.Add(self.appendPassCheckBox)
        self.appendPassCheckBox.CheckedChanged += self.toggle_select_pass_file_button


        # Add Select File button below the passes export append checkbox
        self.selectPassFileButton = Button()
        self.selectPassFileButton.Text = "Select File"
        self.selectPassFileButton.Location = Point(110, 250)  # Adjust the y-coordinate as necessary
        self.selectPassFileButton.Size = Size(90, 25)
        self.selectPassFileButton.Click += lambda sender, args: self.select_file(sender, args, "Pass")
        self.statisticsGroupBox.Controls.Add(self.selectPassFileButton)

        
        # Add append to file checkbox for export failures
        self.appendFailCheckBox = CheckBox()
        self.appendFailCheckBox.Text = "Append to file"
        self.appendFailCheckBox.Location = Point(330, 220)
        self.appendFailCheckBox.Size = Size(100, 25)
        self.statisticsGroupBox.Controls.Add(self.appendFailCheckBox)
        self.appendFailCheckBox.CheckedChanged += self.toggle_select_fail_file_button

        # Add Select File button below the failures export append checkbox
        self.selectFailFileButton = Button()
        self.selectFailFileButton.Text = "Select File"
        self.selectFailFileButton.Location = Point(330, 250)  # Adjust the y-coordinate as necessary
        self.selectFailFileButton.Size = Size(90, 25)
        self.selectFailFileButton.Click += lambda sender, args: self.select_file(sender, args, "Fail")
        self.statisticsGroupBox.Controls.Add(self.selectFailFileButton)

        # Add append to file checkbox for export errors
        self.appendErrorCheckBox = CheckBox()
        self.appendErrorCheckBox.Text = "Append to file"
        self.appendErrorCheckBox.Location = Point(550, 220)
        self.appendErrorCheckBox.Size = Size(100, 25)
        self.statisticsGroupBox.Controls.Add(self.appendErrorCheckBox)
        self.appendErrorCheckBox.CheckedChanged += self.toggle_select_error_file_button

        # Add Select File button below the errors export append checkbox
        self.selectErrorFileButton = Button()
        self.selectErrorFileButton.Text = "Select File"
        self.selectErrorFileButton.Location = Point(550, 250)  # Adjust the y-coordinate as necessary
        self.selectErrorFileButton.Size = Size(90, 25)
        self.selectErrorFileButton.Click += lambda sender, args: self.select_file(sender, args, "Error")
        self.statisticsGroupBox.Controls.Add(self.selectErrorFileButton)

        # Initially set buttons to disabled until checkbox is checked
        self.selectPassFileButton.Enabled = False
        self.selectFailFileButton.Enabled = False
        self.selectErrorFileButton.Enabled = False

        # Create a GroupBox frame for MFGT-RUN
        self.mfgtFrame = GroupBox()
        self.mfgtFrame.Text = "Manufacturing Test Run"
        self.mfgtFrame.Location = Point(20, 500)  # Adjust the location as per your GUI layout. This places it further down the GUI.
        self.mfgtFrame.Size = Size(280, 70)  # This size should comfortably house the button, but you can adjust it if needed.
        self.Controls.Add(self.mfgtFrame)

        # Add the new button for MFGT-RUN within the mfgtFrame GroupBox
        self.bMFGTRun = Button()
        self.bMFGTRun.Text = "MFGT-RUN"
        self.bMFGTRun.Location = Point(30, 20)  # This positions the button within the mfgtFrame GroupBox.
        self.bMFGTRun.Size = Size(200, 30)
        self.bMFGTRun.Click += self.MFGT_RunSequence
        self.mfgtFrame.Controls.Add(self.bMFGTRun)
        self.controlButtonFrame = GroupBox()
        self.controlButtonFrame.Text = "Control Buttons"
        self.controlButtonFrame.Location = Point(20, 60)  # Adjust as necessary
        self.controlButtonFrame.Size = Size(280, 440)  # Adjust the size to encompass all buttons


        # Initialize APx and checkedData
        self.APx = APx500_Application()
        self.checkedData = []


    def APxRunSequence(self, sender, args):
        try:
            self.APx.Sequence.Run()
            self.GetCheckedData(sender, args)  # Call the GetCheckedData after running the sequence
        except Exception as e:
            logging.exception("An error occurred during sequence run:")
            logging.error(f"An unexpected error occurred: {e}\nCheck the log file for more details.")
        
    def updateCheckedMeasurementsList(self, sender, args):
        PASS_COLOR = Color.Green
        FAIL_COLOR = Color.Red
        ERROR_COLOR = Color.Orange

        # Clear the checkedMeasurementsList and checkedResultsList
        self.checkedMeasurementsList.Items.Clear()
        self.checkedResultsList.Items.Clear()  # Clear the results list whenever a new Signal Path is selected

        if self.checkedSignalPathsList.SelectedIndices.Count == 0:
            return

        # Retrieve the selectedSignalPath from self.checkedData using selectedIndex
        selectedIndex = self.checkedSignalPathsList.SelectedIndices[0]
        if selectedIndex < 0 or selectedIndex >= len(self.checkedData):
            logging.error(f"Invalid index selected: {selectedIndex}. Size of checkedData: {len(self.checkedData)}")
            return

        selectedSignalPath = self.checkedData[selectedIndex]

        # Populate the checkedMeasurementsList with red-colored items if they don’t pass the limit checks
        for measurement in selectedSignalPath["measurements"]:
            m_item = ListViewItem(measurement["name"])
            has_error = any(self.result_error(selectedSignalPath["name"], measurement["name"], result["name"]) for result in measurement["results"])
            has_failure = any(not result["passed"] for result in measurement["results"])

            m_item.Font = Font(m_item.Font, FontStyle.Regular)

            if has_error:
                m_item.ForeColor = ERROR_COLOR
            elif has_failure:
                m_item.ForeColor = FAIL_COLOR
            self.checkedMeasurementsList.Items.Add(m_item)
        
        logging.info(f"Size of checkedData: {len(self.checkedData)}")
        logging.info(f"Selected index from checkedSignalPathsList: {selectedIndex}")

    def updateCheckedResultsList(self, sender, args):
        PASS_COLOR = Color.Green
        FAIL_COLOR = Color.Red
        ERROR_COLOR = Color.Orange

        # Clear the checkedResultsList
        self.checkedResultsList.Items.Clear()

        if (self.checkedSignalPathsList.SelectedIndices.Count == 0 or
            self.checkedMeasurementsList.SelectedIndices.Count == 0):
            return

        # Retrieve the selectedSignalPath and selectedMeasurement from self.checkedData using indices
        spIndex = self.checkedSignalPathsList.SelectedIndices[0]

        if spIndex < 0 or spIndex >= len(self.checkedData):
            logging.error(f"Invalid signal path index selected: {spIndex}. Size of checkedData: {len(self.checkedData)}")
            return

        selectedSignalPath = self.checkedData[spIndex]
        mIndex = self.checkedMeasurementsList.SelectedIndices[0]

        if mIndex < 0 or mIndex >= len(selectedSignalPath["measurements"]):
            logging.error(f"Invalid measurement index selected: {mIndex}. Number of measurements in signal path: {len(selectedSignalPath['measurements'])}")
            return

        selectedMeasurement = selectedSignalPath["measurements"][mIndex]

        if spIndex < 0 or spIndex >= len(self.checkedData):
            logging.error(f"Invalid signal path index selected: {spIndex}. Size of checkedData: {len(self.checkedData)}")
            return
        if mIndex < 0 or mIndex >= len(selectedSignalPath["measurements"]):
            logging.error(f"Invalid measurement index selected: {mIndex}. Number of measurements in signal path: {len(selectedSignalPath['measurements'])}")
            return

        # Populate the checkedResultsList with red-colored items if they don’t pass the limit checks
        for result in selectedMeasurement["results"]:
            r_item = ListViewItem(result["name"])
            error = self.result_error(selectedSignalPath["name"], selectedMeasurement["name"], result["name"])
            failed = not result["passed"]

            if error:
                r_item.ForeColor = ERROR_COLOR
            elif failed:
                r_item.ForeColor = FAIL_COLOR


        logging.info(f"Size of checkedData: {len(self.checkedData)}")
        logging.info(f"Selected index from checkedSignalPathsList: {spIndex}")
        logging.info(f"Selected index from checkedMeasurementsList: {mIndex}")

        self.checkedResultsList.Items.Add(r_item)

    def is_result_failed(self, signal_path_name, measurement_name, result_name):
        try:
            for idx in range(self.APx.Sequence.FailedResults.Count):
                failed_result = self.APx.Sequence.FailedResults[idx]
                if (failed_result.Name == result_name and
                    failed_result.MeasurementName == measurement_name and
                    failed_result.SignalPathName == signal_path_name):
                    return not failed_result.PassedResult
        except Exception as e:
            logging.error(f"Error retrieving failed status for result {result_name}: {str(e)}")
        return False
    
    def result_error(self, signal_path_name, measurement_name, result_name):
        try:
            for idx in range(self.APx.Sequence.Results.Count):
                result = self.APx.Sequence.Results[idx]
                if (result.Name == result_name and
                    result.MeasurementName == measurement_name and
                    result.SignalPathName == signal_path_name):
                    return result.HasErrorMessage
        except Exception as e:
            logging.error(f"Error retrieving error status for result {result_name}: {str(e)}")
        return False

    def process_measurement_data(self, sequence_result):
        data = {}
        
        # Check if the result object has the 'PassedLimitChecks' attribute
        if hasattr(sequence_result, 'PassedLimitChecks'):
            passed_limit_checks = sequence_result.PassedLimitChecks
            logging.info(f"\t\tPassedLimitChecks for {sequence_result.Name}: {passed_limit_checks}")
        else:
            passed_limit_checks = False
            logging.warning(f"\t\t{sequence_result.Name} does not have PassedLimitChecks attribute. Setting to False.")

        # Check and Get XYValues
        for vertical_axis in [VerticalAxis.Left, VerticalAxis.Right]:
            channel_count = sequence_result.GetXYChannelCount(vertical_axis)
            if channel_count > 0:
                xValues, yValues = [], []
                for ch in range(channel_count):
                    graphPoints = sequence_result.GetXYValues(ch, vertical_axis, SourceDataType.Measured, 1)
                    if graphPoints:
                        xVals, yVals = [point.X for point in graphPoints], [point.Y for point in graphPoints]
                        xValues.append(xVals)
                        yValues.append(yVals)
                data['xValues'] = xValues
                data['yValues'] = yValues
        
        # Check for meter values
        if sequence_result.HasMeterValues:
            meterValues = sequence_result.GetMeterValues()
            data['meterValues'] = meterValues
            logging.info(f"\t\tFound Meter Values for Result: {sequence_result.Name}")
            # Log each meter value directly.
            for idx, value in enumerate(meterValues):
                logging.info(f"\t\t\tMeter Value {idx}: {value}")

        # Check and Get RawTextResults
        if sequence_result.HasRawTextResults:
            rawTextResults = sequence_result.GetRawTextResults()
            data['rawTextResults'] = rawTextResults
            logging.info(f"\t\tFound Raw Text Results for Result: {sequence_result.Name}")

        # Check and Get XValues and YValues
        for ch in range(sequence_result.GetXYChannelCount(VerticalAxis.Left)):
            for vertical_axis in [VerticalAxis.Left, VerticalAxis.Right]:
                for data_type in [SourceDataType.Measured]:  # Only considering SourceDataType.Measured as per the older logic
                    if hasattr(sequence_result, 'HasXValues') and sequence_result.HasXValues(ch, vertical_axis, data_type):
                        xValues = sequence_result.GetXValues(ch, vertical_axis, data_type, 1)
                        data[f'xValues_{vertical_axis}_{data_type}'] = xValues
                    if hasattr(sequence_result, 'HasYValues') and sequence_result.HasYValues(ch, vertical_axis, data_type):
                        yValues = sequence_result.GetYValues(ch, vertical_axis, data_type, 1)
                        data[f'yValues_{vertical_axis}_{data_type}'] = yValues

        return data

    def GetCheckedData(self, sender=None, args=None):
        PASS_COLOR = Color.Green
        FAIL_COLOR = Color.Red
        ERROR_COLOR = Color.Orange

        if not self.APx:
            logging.error("APx is None. Launch AP Software first.")
            return

        checked_signal_paths = []
        total_results = 0
        failed_results = 0
        error_results = 0  # Added counter for errors
        failed_results_list = []
        error_results_list = []  # Added list to track error results

        try:
            for sp_idx, sp in enumerate(self.APx.Sequence):
                signal_path = ISignalPath(sp)
                logging.info(f"Processing Signal Path: {signal_path.Name}")
                if not signal_path.Checked:
                    continue
                current_sp = {"name": signal_path.Name, "measurements": []}
                logging.info(f"Checked Signal Path: {signal_path.Name}")
                sp_failed, sp_error = False, False  # added sp_error to track signal paths with errors

                for m_idx, m in enumerate(signal_path):
                    measurement = ISequenceMeasurement(m)
                    if not measurement.Checked:
                        continue
                    current_measurement = {"name": measurement.Name, "results": []}
                    logging.info(f"\tChecked Measurement: {measurement.Name}")

                    for result_idx, result in enumerate(measurement.SequenceResults):
                        sequence_result = ISequenceResult(result)
                        failed = self.is_result_failed(signal_path.Name, measurement.Name, sequence_result.Name)
                        error = self.result_error(signal_path.Name, measurement.Name, sequence_result.Name)
                        total_results += 1
                        if failed:
                            failed_results += 1
                            failed_results_list.append((signal_path.Name, measurement.Name, sequence_result.Name))
                            sp_failed = True  # mark signal path as failed
                        if error:
                            error_results += 1  # Increment error count
                            error_results_list.append((signal_path.Name, measurement.Name, sequence_result.Name))
                            sp_error = True  # mark signal path as having an error

                        status = "ERROR" if error else ("FAIL" if failed else "PASS")
                        logging.info(f"{signal_path.Name} | {measurement.Name} | {sequence_result.Name} -> {status}")

                        current_result = {
                            'name': sequence_result.Name,
                            'result_object': sequence_result,
                            'data': self.process_measurement_data(sequence_result),
                            'passed': not (failed or error)
                        }
                        current_measurement["results"].append(current_result)

                    current_sp["measurements"].append(current_measurement)

                checked_signal_paths.append(current_sp)
                logging.info(f"Added a new signal path to checked_signal_paths. Total items now: {len(checked_signal_paths)}")

        except Exception as e:
            logging.error(f"An unexpected error occurred: {e}\nCheck the log file for more details.")

        # Clear and update list boxes
        self.checkedSignalPathsList.Items.Clear()

        for signal_path in checked_signal_paths:
            sp_item = ListViewItem(signal_path["name"])
            has_error = any(self.result_error(signal_path["name"], measurement["name"], m_result["name"]) for measurement in signal_path["measurements"] for m_result in measurement["results"])
            has_failure = any(not m_result["passed"] for measurement in signal_path["measurements"] for m_result in measurement["results"])

            if has_error and has_failure:
                sp_item.ForeColor = FAIL_COLOR
                sp_item.Font = Font(sp_item.Font, FontStyle.Bold)
            elif has_error:
                sp_item.ForeColor = ERROR_COLOR
            elif has_failure:
                sp_item.ForeColor = FAIL_COLOR

            self.checkedSignalPathsList.Items.Add(sp_item)
            logging.info(f"Added a new item to checkedSignalPathsList. Total items now: {self.checkedSignalPathsList.Items.Count}")
        
        # Update the self.checkedData based on the populated checked_signal_paths
        self.checkedData = checked_signal_paths
        self.total_tests_cumulative += total_results
        self.total_passed_cumulative += (total_results - failed_results - error_results)
        self.total_failed_cumulative += failed_results
        self.total_errors_cumulative += error_results

        self.update_statistics(self.total_tests_cumulative, self.total_failed_cumulative, self.total_errors_cumulative, failed_results_list, error_results_list, checked_signal_paths)

        return checked_signal_paths
    
    def update_statistics(self, total_results, failed_results, error_results, failed_results_list, error_results_list, checked_signal_paths):
        PASS_COLOR = Color.Green
        FAIL_COLOR = Color.Red
        ERROR_COLOR = Color.Orange

        # Displaying rates
        pass_rate = ((total_results - (failed_results + error_results)) / total_results) * 100 if total_results else 0
        fail_rate = (failed_results / total_results) * 100 if total_results else 0
        error_rate = (error_results / total_results) * 100 if total_results else 0  # Calculate error rate

        logging.info(f"Total test results: {total_results}")
        logging.info(f"Number of failures: {failed_results}")
        logging.info(f"Number of errors: {error_results}")  
        logging.info(f"Pass rate: {pass_rate:.2f}%")
        logging.info(f"Fail rate: {fail_rate:.2f}%")
        logging.info(f"Error rate: {error_rate:.2f}%")  
        logging.info(f"List of measurement results that failed: {', '.join([f'{s} | {m} | {r}' for s, m, r in failed_results_list])}")
        logging.info(f"List of measurement results that errored: {', '.join([f'{s} | {m} | {r}' for s, m, r in error_results_list])}")  

        self.passRateLabel.Text = f"Pass Rate: {pass_rate:.2f}%"
        self.failRateLabel.Text = f"Fail Rate: {fail_rate:.2f}%"
        self.errorRateLabel.Text = f"Error Rate: {error_rate:.2f}%" 
            
        # For displaying Number of Passed, Failed, and Errors
        self.passedNumberLabel.Text = f"{total_results - failed_results - error_results}"
        self.failedNumberLabel.Text = f"{failed_results}"
        self.errorNumberLabel.Text = f"{error_results}"

        # For updating Number of Items
        self.totalItemsLabel.Text = f"Total # of Tests: {total_results}"

        # Updating List of Passing, Failing, and Errors
        passing_list = [f"{sp['name']} | {m['name']} | {r['name']}" 
                        for sp in checked_signal_paths 
                        for m in sp['measurements'] 
                        for r in m['results'] 
                        if not (self.is_result_failed(sp['name'], m['name'], r['name']) or self.result_error(sp['name'], m['name'], r['name']))]
        
        self.passedListBox.Text = "List of Passing: " + '\n'.join(passing_list) + '\n'
        self.failureListTextBox.Text = "List of Failures: " + '\n'.join([f"{s} | {m} | {r}" for s, m, r in failed_results_list]) + '\n'
        self.errorListTextBox.Text = "List of Errors: " + '\n'.join([f"{s} | {m} | {r}" for s, m, r in error_results_list]) + '\n'

    def ClearAll(self, sender, args):
        self.checkedSignalPathsList.Items.Clear()
        self.checkedMeasurementsList.Items.Clear()
        self.checkedResultsList.Items.Clear()
        self.selectedResultsList.Items.Clear()
        self.passedListBox.Text = ""
        self.failureListTextBox.Text = ""
        self.errorListTextBox.Text = ""
        self.passRateLabel.Text = ""
        self.failRateLabel.Text = ""
        self.errorRateLabel.Text = ""
        self.passedNumberLabel.Text = ""
        self.failedNumberLabel.Text = ""
        self.errorNumberLabel.Text = ""
        self.totalItemsLabel.Text = ""


    def ExportCheckedDataToExcel(self, sender=None, args=None):
        # Check if there is any checked data to export
        if not self.checkedData:
            logging.warning("No checked data to export.")
            return

        # If "Append to" is checked and a file is selected, then append to that file
        if self.appendCheckbox.Checked and hasattr(self, "selectedAppendFilePath"):
            self.append_to_existing_excel(self.selectedAppendFilePath, self.checkedData, self.unitInput.Text.strip())
        else:
            # Call the export_to_excel method with the checked data
            self.export_to_excel(self.checkedData, self.unitInput.Text.strip())


    @staticmethod
    def sanitize_sheet_name(name):
        invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '-')
        return name[:31]

    @staticmethod
    def abbreviate_name(name, max_length=10):
        if len(name) <= max_length:
            return name
        words = name.split(' ')
        if len(words) == 1:
            return name[:max_length]
        abbreviation = ''.join([word[0] for word in words])
        if len(abbreviation) <= max_length:
            return abbreviation
        return abbreviation[:max_length]
    
    def unique_sheet_name(self, workbook, base_name):
        count = 2
        sheet_name = base_name
        while sheet_name in workbook.sheetnames:
            sheet_name = f"{base_name} ({count})"
            count += 1
        return sheet_name

    def export_to_excel(self, checked_signal_paths, unit_descriptor=None, args=None):
        wb = Workbook()
        ws = wb.active
        has_created_sheet = False

        try:
            # unit_no = self.unitInput.Text.strip()
            if unit_descriptor:  # Incorporate the descriptor into the filename
                file_name = f"{unit_descriptor}.xlsx" if unit_descriptor else "exported_data.xlsx"

            # Define cleaned_descriptor to remove the suffixes
            cleaned_descriptor = unit_descriptor.replace("_PASS", "").replace("_FAIL", "").replace("_ERRORS", "")

            for sp in checked_signal_paths:
                for measurement in sp["measurements"]:
                    for result in measurement["results"]:

                        xy_values_found = False  # Add a flag to track if xy values are found

                        if 'xValues' in result['data'] and 'yValues' in result['data']:
                            # Ensure xValues is always a list.
                            xValues = result['data']['xValues']
                            if isinstance(xValues[0], (list, tuple)):
                                xValues = xValues[0]
                            
                            sheet_title = f"{self.abbreviate_name(sp['name'])}_{self.abbreviate_name(measurement['name'])}_{self.abbreviate_name(result['name'])}"
                            sheet_name = self.unique_sheet_name(wb, sheet_title)
                            ws = wb.create_sheet(title=sheet_name)
                            ws.append([f'Signal Path: {sp["name"]}'])
                            ws.append([f'Measurement: {measurement["name"]}'])
                            ws.append([f'Result: {result["name"]}'])
                            # Modified header to include Serial Number for Y Values
                            cleaned_descriptor = unit_descriptor.replace("_PASS", "").replace("_FAIL", "").replace("_ERRORS", "")
                            ws.append(["X Values"] + [f"{cleaned_descriptor} Ch{idx+1}" for idx in range(len(result['data']['yValues']))])
                            if "ch2" in result['name'].lower():
                                continue  # skip the ch2 result to avoid duplicating data
                        
                            # Writing to Excel
                            for row_idx, xValue in enumerate(xValues):
                                yRow = []
                                for ch_idx, yValueSet in enumerate(result['data']['yValues']):
                                    if not isinstance(yValueSet, (list, tuple)):
                                        logging.error(f"Invalid yValue data structure at yValues channel index {ch_idx}. Expected list or tuple, but got {type(yValueSet)}.")
                                        yValueSet = [yValueSet]  # Convert single float to a list for consistency.
                                    
                                    try:
                                        yRow.append(yValueSet[row_idx])
                                    except IndexError:
                                        logging.warning(f"Index {row_idx} out of bounds for yValues set at channel index {ch_idx}.")
                                        yRow.append(None)
                                row = [xValue] + yRow
                                ws.append(row)
                                has_created_sheet = True 
                                xy_values_found = True  # Set the flag here!


                        if 'meterValues' in result['data']:
                            # Create a new sheet for Meter Values.
                            meter_sheet_title = f"{self.abbreviate_name(sp['name'])}_{self.abbreviate_name(measurement['name'])}_{self.abbreviate_name(result['name'])}"
                            meter_sheet_name = self.unique_sheet_name(wb, meter_sheet_title)
                            meter_ws = wb.create_sheet(title=meter_sheet_name)
                            
                            meter_ws.append([f'Signal Path: {sp["name"]}'])
                            meter_ws.append([f'Measurement: {measurement["name"]}'])
                            meter_ws.append([f'Result: {result["name"]}'])
                            
                            # Add the headers for 'Channels' and the serial number
                            meter_ws.append(['Channels', cleaned_descriptor])
                            
                            meterValues = result['data']['meterValues']
                            
                            # Adjusted to add the channel name to the start of each row
                            for idx, val in enumerate(meterValues):
                                meter_ws.append([f"Ch{idx+1}", val])  # The channel name "Chx" and its corresponding value on the same row
                                                        
                            has_created_sheet = True  # Set has_created_sheet to True when creating a sheet for meterValues
                                        
                        if 'rawTextResults' in result['data']:
                            ws.append(["Raw Text Results"])
                            rawTextResults = result['data']['rawTextResults']
                            ws.append(rawTextResults)
                            has_created_sheet = True

                        for vertical_axis in [VerticalAxis.Left, VerticalAxis.Right]:
                            for data_type_str in ["Measured", "Fitted", "Residual"]:
                                if hasattr(SourceDataType, data_type_str):
                                    data_type = getattr(SourceDataType, data_type_str)
                                x_key = f'xValues_{vertical_axis}_{data_type}'
                                y_key = f'yValues_{vertical_axis}_{data_type}'
                                if x_key in result['data'] and y_key in result['data']:
                                    ws.append([f"X Values ({vertical_axis}, {data_type})"] + [f"Y Values CH{idx+1} ({vertical_axis}, {data_type})" for idx in range(len(result['data'][y_key]))])
                                    for row_idx, xValue in enumerate(result['data'][x_key]):
                                        row = [xValue] + [result['data'][y_key][ch_idx][row_idx] for ch_idx in range(len(result['data'][y_key]))]
                                        ws.append(row)
                                    xy_values_found = True  # Update the flag if xy values are found

                        if not xy_values_found:  # This remains outside the nested loop
                            logging.warning(f"{result['name']} does not have xy values.")

            if has_created_sheet:
                # Remove the default "Sheet" if it exists in the workbook.
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
                wb.save(file_name)
                logging.info(f"Exported data to {file_name}.")
            else:
                logging.warning("No data to export.")
        except Exception as e:
            logging.exception("An error occurred during Excel export:")
            logging.error(f"An unexpected error occurred: {e}\nCheck the log file for more details.")
    
    def filter_data(self, condition):
        filtered_data = []
        for sp in self.checkedData:
            new_sp = dict(sp)  # Create a shallow copy of the signal path dict.
            new_measurements = []
            for m in sp['measurements']:
                new_m = dict(m)  # Create a shallow copy of the measurement dict.
                new_results = [r for r in m['results'] if condition(sp['name'], m['name'], r['name'])]
                if new_results:
                    new_m['results'] = new_results
                    new_measurements.append(new_m)
            if new_measurements:
                new_sp['measurements'] = new_measurements
                filtered_data.append(new_sp)
        return filtered_data

    def export_pass(self, sender=None, args=None):
        passing_data = self.filter_data(lambda sp, m, r: not (self.is_result_failed(sp, m, r) or self.result_error(sp, m, r)))
        if self.appendPassCheckBox.Checked and self.selectedPassFilePath:
            self.append_to_existing_excel(self.selectedPassFilePath, passing_data, self.unitInput.Text.strip() + "_PASS")
        else:
            self.export_to_excel(passing_data, self.unitInput.Text.strip() + "_PASS")

    def export_fail(self, sender=None, args=None):
        failed_data = self.filter_data(self.is_result_failed)
        if self.appendFailCheckBox.Checked and self.selectedFailFilePath:
            self.append_to_existing_excel(self.selectedFailFilePath, failed_data, self.unitInput.Text.strip() + "_FAIL")
        else:
            self.export_to_excel(failed_data, self.unitInput.Text.strip() + "_FAIL")

    def export_error(self, sender=None, args=None):
        error_data = self.filter_data(self.result_error)
        if self.appendErrorCheckBox.Checked and self.selectedErrorFilePath:
            self.append_to_existing_excel(self.selectedErrorFilePath, error_data, self.unitInput.Text.strip() + "_ERRORS")
        else:
            self.export_to_excel(error_data, self.unitInput.Text.strip() + "_ERRORS")
    def AddSelectedResult(self, sender, args):
        # Get the selected items in the checkedResultsList ListBox.
        selected_items = self.checkedResultsList.SelectedItems

        # Check if there are selected indices in the signal paths and measurements list
        if (self.checkedSignalPathsList.SelectedIndices.Count == 0 or
            self.checkedMeasurementsList.SelectedIndices.Count == 0):
            logging.warning("Signal Path or Measurement is not selected.")
            return

        # Retrieve the currently selected signal path and measurement indices
        spIndex = self.checkedSignalPathsList.SelectedIndices[0]
        mIndex = self.checkedMeasurementsList.SelectedIndices[0]

        # Retrieve the names of the selected signal path and measurement
        signal_path_name = self.checkedSignalPathsList.Items[spIndex].Text
        measurement_name = self.checkedMeasurementsList.Items[mIndex].Text

        # Iterate over the selected items in the results list and form the full names
        for item in selected_items:
            full_name = f"{signal_path_name} | {measurement_name} | {item.Text}"
            if not self.is_duplicate(full_name):
                self.selectedResultsList.Items.Add(full_name)

    def PinSelectedSignalPath(self, sender, args):
        # Check if there is a selected index in the signal paths list
        if self.checkedSignalPathsList.SelectedIndices.Count == 0:
            logging.warning("Signal Path is not selected.")
            return

        # Loop through each selected signal path
        for spIndex in self.checkedSignalPathsList.SelectedIndices:
            selectedSignalPath = self.checkedData[spIndex]
            signal_path_name = selectedSignalPath["name"]

            # Iterate over all the measurements in the selectedSignalPath
            for measurement in selectedSignalPath["measurements"]:
                measurement_name = measurement["name"]

                # Now, iterate over all the results associated with the current measurement
                for result in measurement["results"]:
                    result_name = result["name"]

                    # Form the full name and add it to the selectedResultsList ListBox
                    full_name = f"{signal_path_name} | {measurement_name} | {result_name}"
                    if not self.is_duplicate(full_name):
                        self.selectedResultsList.Items.Add(full_name)
    
    def PinSelectedMeasurement(self, sender, args):
        # Check if there are selected items in the signal paths and measurements lists
        if self.checkedSignalPathsList.SelectedIndices.Count == 0 or self.checkedMeasurementsList.SelectedIndices.Count == 0:
            logging.warning("Signal Path or Measurement is not selected.")
            return

        # Loop through each selected signal path
        for spIndex in self.checkedSignalPathsList.SelectedIndices:
            selectedSignalPath = self.checkedData[spIndex]
            signal_path_name = selectedSignalPath["name"]

            # Loop through each selected measurement for the current signal path
            for mIndex in self.checkedMeasurementsList.SelectedIndices:
                selectedMeasurement = selectedSignalPath["measurements"][mIndex]
                measurement_name = selectedMeasurement["name"]

                # Now, iterate over all the results associated with the selected measurement
                for result in selectedMeasurement["results"]:
                    result_name = result["name"]

                    # Form the full name and add it to the selectedResultsList ListBox
                    full_name = f"{signal_path_name} | {measurement_name} | {result_name}"
                    if not self.is_duplicate(full_name):
                        self.selectedResultsList.Items.Add(full_name)

    def is_duplicate(self, full_name):
        for item in self.selectedResultsList.Items:
            if item == full_name:
                return True
        return False
    
    def RerunPinnedTests(self, sender, args):
        # Create a dictionary where keys are signal paths and values are sets of measurements
        signal_path_to_measurements = {}

        # Populate the dictionary from the selectedResultsList
        for item in self.selectedResultsList.Items:
            # Split the full name into its components
            components = item.split(' | ')
            signal_path = components[0]
            measurement = components[1]

            # Add the measurement to the set of measurements for this signal path
            if signal_path not in signal_path_to_measurements:
                signal_path_to_measurements[signal_path] = set()
            signal_path_to_measurements[signal_path].add(measurement)

        # Check which measurements to rerun
        for signal_path, measurements in signal_path_to_measurements.items():
            selectedSignalPath = next(sp for sp in self.checkedData if sp["name"] == signal_path)
            if len(measurements) == len(selectedSignalPath["measurements"]):
                # All measurements for this signal path are selected, rerun the entire signal path
                self.rerun_signal_path_in_AP(signal_path)
            else:
                # Only specific measurements are selected, rerun those
                for measurement in measurements:
                    self.rerun_measurement_in_AP(signal_path, measurement)

    def rerun_signal_path_in_AP(self, signal_path):
        # TODO: Add the implementation to rerun an entire signal path in APx software
        pass

    def rerun_measurement_in_AP(self, signal_path, measurement):
        # TODO: Add the implementation to rerun a specific measurement under a given signal path in APx software
        pass

    def ExportSelectedResults(self, sender, args):
        logging.info("Entered ExportSelectedResults function.")  # Log Entry Point
        
        logging.info(f"Number of items in selectedResultsList: {self.selectedResultsList.Items.Count}")

        selected_items = [self.selectedResultsList.Items[i] for i in range(self.selectedResultsList.Items.Count)]

        if not selected_items:
            logging.warning("No data to export.")
            return

        logging.info(f"Selected items: {selected_items}")

        selected_signal_paths = []

        if not self.checkedData:
            logging.warning("checkedData is empty.")
            return

        for sp in self.checkedData:
            sp_name = sp.get("name")
            if not sp_name:
                logging.warning(f"Signal path does not have a name: {sp}")
                continue

            copied_sp = {'name': sp.get('name'), 'measurements': [], 'meterValues': sp.get('meterValues', [])}

            for measurement in sp.get("measurements", []):
                measurement_name = measurement.get("name")
                if not measurement_name:
                    logging.warning(f"Measurement does not have a name: {measurement}")
                    continue

                copied_measurement = {'name': measurement_name, 'results': []}
                logging.info(f"Copied Measurement: {copied_measurement}")

                for result in measurement.get("results", []):
                    result_name = result.get("name")
                    if not result_name:
                        logging.warning(f"Result does not have a name: {result}")
                        continue

                    full_name = f"{sp_name} | {measurement_name} | {result_name}"
                    logging.info(f"Formed full name: {full_name}")

                    if full_name in selected_items:
                        copied_result = {'name': result_name, 'result_object': result.get('result_object'), 'data': {}}
                        
                        if 'meterValues' in result.get('data', {}):
                            # Convert the .NET array to a Python list
                            meter_values_list = list(result['data']['meterValues'])
                            copied_result['data']['meterValues'] = meter_values_list
                            logging.info(f"Meter Values: {meter_values_list}")
                        
                        if 'xValues' in result['data'] or 'yValues' in result['data']:
                            if 'xValues' in result['data'] and 'yValues' in result['data']:
                                copied_result['data']['xValues'] = result['data']['xValues']
                                copied_result['data']['yValues'] = result['data']['yValues']
                            else:
                                logging.warning(f"Either xValues or yValues is missing from result: {result_name} in measurement: {measurement_name} of signal path: {sp_name}")
                                continue  # Skip this result and move to the next one
                        
                        if copied_result['data']:
                            copied_measurement['results'].append(copied_result)

                if copied_measurement['results']:
                    copied_sp['measurements'].append(copied_measurement)

            if copied_sp['measurements'] or copied_sp['meterValues']:
                selected_signal_paths.append(copied_sp)

        #logging.info(f"Selected signal paths: {selected_signal_paths}")

        if not selected_signal_paths:
            logging.warning("No matched data to export.")
            return

        try:
            self.export_to_excel(selected_signal_paths, self.unitInput.Text.strip())
        except Exception as e:
            logging.exception("Error in calling export_to_excel")
            logging.error(f"An unexpected error occurred: {e}")
        
        logging.info("Exited ExportSelectedResults function.")  # Log Exit Point

    def select_file(self, sender, args, category):
        root = tk.Tk()  # create a root window
        root.withdraw()  # hide the root window
        # only allow excel files
        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls;*.xlsx")])

        if filepath:
            filename = os.path.basename(filepath)  # get the file name without the path
            if category == "Pass":
                self.selectedPassFilePath = filepath
                self.selectPassFileButton.Text = filename
                self.selectPassFileButton.BackColor = Color.Green
            elif category == "Fail":
                self.selectedFailFilePath = filepath
                self.selectFailFileButton.Text = filename
                self.selectFailFileButton.BackColor = Color.Green
            elif category == "Error":
                self.selectedErrorFilePath = filepath
                self.selectErrorFileButton.Text = filename
                self.selectErrorFileButton.BackColor = Color.Green
            elif category == "All":
                self.selectedAppendFilePath = filepath
                self.bSelectFile.Text = filename
                self.bSelectFile.BackColor = Color.Green

    def append_to_existing_excel(self, filename, checked_signal_paths, unit_descriptor):
        logging.info("Starting the appending process...")
        try:
            wb = load_workbook(filename)
        except Exception as e:
            logging.error(f"Failed to load the workbook '{filename}'. Error: {e}")
            return

        unit_no = self.unitInput.Text.strip()
        
        # Define cleaned_descriptor to remove the suffixes
        cleaned_descriptor = unit_descriptor.replace("_PASS", "").replace("_FAIL", "").replace("_ERRORS", "")

        for sp_idx, sp in enumerate(checked_signal_paths):
            logging.info(f"Processing Signal Path {sp_idx + 1}: {sp['name']}")
            for meas_idx, measurement in enumerate(sp["measurements"]):
                for res_idx, result in enumerate(measurement["results"]):
                    sheet_title = f"{self.abbreviate_name(sp['name'])}_{self.abbreviate_name(measurement['name'])}_{self.abbreviate_name(result['name'])}"

                    # Handle existing sheet or create a new one
                    if sheet_title in wb.sheetnames:
                        ws = wb[sheet_title]
                    else:
                        ws = wb.create_sheet(title=self.sanitize_sheet_name(sheet_title))
                        ws.append([f'Signal Path: {sp["name"]}'])
                        ws.append([f'Measurement: {measurement["name"]}'])
                        ws.append([f'Result: {result["name"]}'])

                    # Handle xy values
                    if 'xValues' in result['data'] and 'yValues' in result['data']:
                        xValues = result['data']['xValues']
                        if isinstance(xValues[0], (list, tuple)):
                            xValues = xValues[0]

                        yValues = result['data']['yValues']
                        if not isinstance(yValues[0], (list, tuple)):
                            yValues = [yValues]

                        col = ws.max_column + 1
                        for channel_idx, y_set in enumerate(yValues):
                            # Adjust the header name to follow "<Serial number> <CHx>" convention
                            ws.cell(row=4, column=col, value=f"{cleaned_descriptor} Ch{channel_idx+1}")  # <-- Adjusted this line to fix naming convention
                            for row_idx, y_val in enumerate(y_set):
                                ws.cell(row=row_idx + 5, column=col, value=y_val)
                            col += 1


                    # Handle meter values
                    if 'meterValues' in result['data']:
                        meter_sheet_title = f"{self.abbreviate_name(sp['name'])}_{self.abbreviate_name(measurement['name'])}_{self.abbreviate_name(result['name'])}"

                        if meter_sheet_title in wb.sheetnames:
                            meter_ws = wb[meter_sheet_title]
                        else:
                            meter_ws = wb.create_sheet(title=meter_sheet_title)
                            meter_ws.append([f'Signal Path: {sp["name"]}'])
                            meter_ws.append([f'Measurement: {measurement["name"]}'])
                            meter_ws.append([f'Result: {result["name"]}'])
                            meter_ws.append([unit_descriptor])

                        last_column = meter_ws.max_column + 1
                        meter_ws.cell(row=4, column=last_column, value=unit_no)
                        for idx, val in enumerate(result['data']['meterValues'], start=5):
                            meter_ws.cell(row=idx, column=last_column, value=val)
                        logging.info(f"Appended meter values to sheet: {meter_ws.title}")

                    # Handle raw text results
                    if 'rawTextResults' in result['data']:
                        rawTextResults = result['data']['rawTextResults']
                        for item in rawTextResults:
                            ws.append([item])

        wb.save(filename)
        logging.info(f"Data successfully appended to {filename}")

    def toggle_select_pass_file_button(self, sender, args):
        self.selectPassFileButton.Enabled = self.appendPassCheckBox.Checked

    def toggle_select_fail_file_button(self, sender, args):
        self.selectFailFileButton.Enabled = self.appendFailCheckBox.Checked

    def toggle_select_error_file_button(self, sender, args):
        self.selectErrorFileButton.Enabled = self.appendErrorCheckBox.Checked

    def toggleSelectFileButton(self, sender, args):
        self.bSelectFile.Enabled = self.appendCheckbox.Checked

    def MFGT_RunSequence(self, sender=None, args=None):
        self.APxRunSequence(sender, args)  # Run sequence
        self.GetCheckedData()  # Make sure to retrieve and update stats after running the sequence
        self.ExportCheckedDataToExcel()  # Exports all data to excel
        self.export_pass()  # Export pass data
        self.export_fail()  # Export fail data
        self.export_error()  # Export error data

    def reset_cumulative_data(self):
        self.total_tests_cumulative = 0
        self.total_passed_cumulative = 0
        self.total_failed_cumulative = 0
        self.total_errors_cumulative = 0


def main():
    Application.EnableVisualStyles()
    form = APxContainer()
    Application.Run(form)

if __name__ == "__main__":
    main()